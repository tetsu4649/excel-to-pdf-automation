#!/usr/bin/env python3
"""
新機能のテストスクリプト
- .xlsmファイル対応
- テキストのみPDF出力
- 列選択機能
"""

import os
from pathlib import Path
from excel_to_pdf import ExcelToWordPDFConverter
from openpyxl import Workbook

def create_test_xlsm():
    """テスト用の.xlsmファイルを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "テストデータ"
    
    # ヘッダー
    headers = ["A列", "B列", "C列", "D列", "E列"]
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=header)
    
    # データ
    data_rows = [
        ["製品1", "100", "東京", "2024-01-01", "完了"],
        ["製品2", "200", "大阪", "2024-01-02", "進行中"],
        ["製品3", "300", "名古屋", "2024-01-03", "保留"],
        ["製品4", "400", "福岡", "2024-01-04", "完了"],
        ["製品5", "500", "札幌", "2024-01-05", "進行中"],
    ]
    
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # .xlsmとして保存（実際のマクロは含まれない）
    wb.save("test_data.xlsm")
    print("✅ test_data.xlsm を作成しました")
    return "test_data.xlsm"

def test_xlsm_support():
    """xlsmファイルのサポートをテスト"""
    print("\n=== xlsmファイルサポートのテスト ===")
    
    xlsm_file = create_test_xlsm()
    converter = ExcelToWordPDFConverter()
    
    try:
        sheets = converter.get_sheet_names(xlsm_file)
        print(f"シート一覧: {sheets}")
        
        word_path, pdf_path = converter.convert(xlsm_file, None, sheets[0])
        print(f"✅ xlsmファイルの変換成功: {pdf_path}")
    except Exception as e:
        print(f"❌ エラー: {e}")

def test_text_only_mode():
    """テキストのみモードのテスト"""
    print("\n=== テキストのみモードのテスト ===")
    
    # 通常モード
    converter_normal = ExcelToWordPDFConverter(text_only=False)
    word_path, pdf_path = converter_normal.convert("test_data.xlsm", None, "テストデータ")
    os.rename(pdf_path, "test_normal_mode.pdf")
    print(f"✅ 通常モードPDF: test_normal_mode.pdf")
    
    # テキストのみモード
    converter_text = ExcelToWordPDFConverter(text_only=True)
    word_path, pdf_path = converter_text.convert("test_data.xlsm", None, "テストデータ")
    os.rename(pdf_path, "test_text_only_mode.pdf")
    print(f"✅ テキストのみモードPDF: test_text_only_mode.pdf")

def test_column_selection():
    """列選択機能のテスト"""
    print("\n=== 列選択機能のテスト ===")
    
    # B列のみ（デフォルト）
    converter_b = ExcelToWordPDFConverter(text_only=True, selected_columns=["B"])
    word_path, pdf_path = converter_b.convert("test_data.xlsm", None, "テストデータ")
    os.rename(pdf_path, "test_column_B_only.pdf")
    print(f"✅ B列のみ: test_column_B_only.pdf")
    
    # 複数列選択（A, B, D列）
    converter_multi = ExcelToWordPDFConverter(text_only=True, selected_columns=["A", "B", "D"])
    word_path, pdf_path = converter_multi.convert("test_data.xlsm", None, "テストデータ")
    os.rename(pdf_path, "test_columns_A_B_D.pdf")
    print(f"✅ A,B,D列: test_columns_A_B_D.pdf")
    
    # 全列
    converter_all = ExcelToWordPDFConverter(text_only=True, selected_columns=["ALL"])
    word_path, pdf_path = converter_all.convert("test_data.xlsm", None, "テストデータ")
    os.rename(pdf_path, "test_all_columns.pdf")
    print(f"✅ 全列: test_all_columns.pdf")

def main():
    """メイン関数"""
    print("新機能テストを開始します...\n")
    
    # 出力ディレクトリを作成
    output_dir = Path("test_output")
    output_dir.mkdir(exist_ok=True)
    os.chdir(output_dir)
    
    try:
        # 各機能をテスト
        test_xlsm_support()
        test_text_only_mode()
        test_column_selection()
        
        print("\n✅ すべてのテストが完了しました！")
        print(f"出力ファイルは {output_dir} ディレクトリに保存されています。")
        
    except Exception as e:
        print(f"\n❌ テスト中にエラーが発生しました: {e}")
    
    finally:
        # クリーンアップ
        os.chdir("..")

if __name__ == "__main__":
    main()