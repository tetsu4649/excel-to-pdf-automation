#!/usr/bin/env python3
"""
サンプルExcelファイルを作成するスクリプト
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


def create_sample_excel():
    """テスト用のサンプルExcelファイルを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "売上データ"
    
    # ヘッダー行
    headers = ["日付", "商品名", "数量", "単価", "売上金額", "担当者"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # サンプルデータ
    sample_data = [
        ["2024/01/10", "ノートPC", 5, 120000, 600000, "田中"],
        ["2024/01/12", "マウス", 20, 3000, 60000, "佐藤"],
        ["2024/01/15", "キーボード", 15, 8000, 120000, "鈴木"],
        ["2024/01/18", "モニター", 8, 35000, 280000, "田中"],
        ["2024/01/20", "USBメモリ", 50, 2000, 100000, "高橋"],
        ["2024/01/22", "Webカメラ", 12, 5000, 60000, "佐藤"],
        ["2024/01/25", "ヘッドセット", 25, 4000, 100000, "鈴木"],
        ["2024/01/28", "外付けHDD", 10, 12000, 120000, "田中"],
        ["2024/01/30", "プリンター", 3, 45000, 135000, "高橋"],
        ["2024/02/02", "スキャナー", 5, 25000, 125000, "佐藤"],
    ]
    
    # データを書き込む
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in [3, 4, 5]:  # 数値列は右寄せ
                cell.alignment = Alignment(horizontal="right")
    
    # 列幅を調整
    column_widths = [12, 15, 8, 10, 12, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 合計行を追加
    total_row = len(sample_data) + 2
    ws.cell(row=total_row, column=1, value="合計")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    # 数量の合計
    total_quantity = sum(row[2] for row in sample_data)
    ws.cell(row=total_row, column=3, value=total_quantity)
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    
    # 売上金額の合計
    total_sales = sum(row[4] for row in sample_data)
    ws.cell(row=total_row, column=5, value=total_sales)
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    
    # ファイルを保存
    wb.save("sample_data.xlsx")
    print("サンプルExcelファイル 'sample_data.xlsx' を作成しました。")
    
    # 別のシートも追加
    ws2 = wb.create_sheet("在庫管理")
    
    # 在庫管理シートのヘッダー
    inventory_headers = ["商品コード", "商品名", "在庫数", "最小在庫数", "状態"]
    for col, header in enumerate(inventory_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # 在庫データ
    inventory_data = [
        ["PC001", "ノートPC", 25, 10, "正常"],
        ["MS001", "マウス", 150, 50, "正常"],
        ["KB001", "キーボード", 45, 30, "正常"],
        ["MN001", "モニター", 8, 15, "要発注"],
        ["USB001", "USBメモリ", 200, 100, "正常"],
        ["WC001", "Webカメラ", 5, 20, "要発注"],
        ["HS001", "ヘッドセット", 80, 40, "正常"],
        ["HDD001", "外付けHDD", 35, 20, "正常"],
    ]
    
    # 在庫データを書き込む
    for row_idx, row_data in enumerate(inventory_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            # 状態が"要発注"の場合は赤色にする
            if col_idx == 5 and value == "要発注":
                cell.font = Font(color="FF0000", bold=True)
    
    # 再度保存
    wb.save("sample_data.xlsx")
    
    return "sample_data.xlsx"


if __name__ == "__main__":
    create_sample_excel()